"""
Reports API endpoints
"""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, and_, or_
from datetime import datetime, timedelta, date
from typing import Optional, List, Tuple
from enum import Enum
from io import BytesIO
import logging

from app.core.database import get_db
from app.models.user import User
from app.models.project import Project
from app.models.building import Building
from app.models.owner import Owner
from app.models.interaction import Interaction
from app.models.unit import Unit
from app.models.task import Task
from app.models.document import DocumentSignature
from app.api.dependencies import get_current_user
from pydantic import BaseModel

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/reports", tags=["reports"])


class ReportType(str, Enum):
    """Available report types"""
    BUILDING_PROGRESS = "building_progress"
    AGENT_PERFORMANCE = "agent_performance"
    INTERACTION_HISTORY = "interaction_history"
    COMPLIANCE_AUDIT = "compliance_audit"


class ReportFormat(str, Enum):
    """Available export formats"""
    PDF = "pdf"
    EXCEL = "excel"


class ReportRequest(BaseModel):
    """Report generation request"""
    report_type: ReportType
    format: ReportFormat = ReportFormat.PDF
    project_id: Optional[str] = None
    building_id: Optional[str] = None
    agent_id: Optional[str] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None


@router.get("/types")
async def get_report_types(
    current_user: User = Depends(get_current_user)
):
    """Get list of available report types"""
    return {
        "report_types": [
            {
                "id": "building_progress",
                "name": "Building Progress Report",
                "description": "Shows signature progress and status for buildings",
                "formats": ["pdf", "excel"]
            },
            {
                "id": "agent_performance",
                "name": "Agent Performance Report",
                "description": "Shows agent KPIs and interaction statistics",
                "formats": ["pdf", "excel"]
            },
            {
                "id": "interaction_history",
                "name": "Interaction History Report",
                "description": "Detailed log of all interactions with owners",
                "formats": ["pdf", "excel"]
            },
            {
                "id": "compliance_audit",
                "name": "Compliance Audit Report",
                "description": "Audit trail of all approvals and signatures",
                "formats": ["pdf", "excel"]
            }
        ]
    }


@router.post("/generate")
async def generate_report(
    request: ReportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate a report in the specified format
    
    Returns the generated file as a download
    """
    try:
        if request.format == ReportFormat.PDF:
            file_content, filename = await generate_pdf_report(
                request, db, current_user
            )
            return StreamingResponse(
                BytesIO(file_content),
                media_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )
        elif request.format == ReportFormat.EXCEL:
            file_content, filename = await generate_excel_report(
                request, db, current_user
            )
            return StreamingResponse(
                BytesIO(file_content),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )
        else:
            raise HTTPException(status_code=400, detail="Unsupported format")
    except Exception as e:
        logger.error(f"Error generating report: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")


async def generate_pdf_report(
    request: ReportRequest,
    db: Session,
    current_user: User
) -> Tuple[bytes, str]:
    """Generate PDF report"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#0D9488'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Header
        title_text = request.report_type.value.replace('_', ' ').title() + " Report"
        story.append(Paragraph(title_text, title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Report metadata
        metadata_style = ParagraphStyle(
            'Metadata',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey
        )
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", metadata_style))
        story.append(Paragraph(f"Generated by: {current_user.full_name or current_user.email}", metadata_style))
        if request.start_date or request.end_date:
            date_range = f"{request.start_date or 'Start'} to {request.end_date or 'End'}"
            story.append(Paragraph(f"Date Range: {date_range}", metadata_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Generate report content based on type
        if request.report_type == ReportType.BUILDING_PROGRESS:
            content = await _get_building_progress_data(request, db, current_user)
            story.extend(_build_building_progress_pdf(content, styles))
        elif request.report_type == ReportType.AGENT_PERFORMANCE:
            content = await _get_agent_performance_data(request, db, current_user)
            story.extend(_build_agent_performance_pdf(content, styles))
        elif request.report_type == ReportType.INTERACTION_HISTORY:
            content = await _get_interaction_history_data(request, db, current_user)
            story.extend(_build_interaction_history_pdf(content, styles))
        elif request.report_type == ReportType.COMPLIANCE_AUDIT:
            content = await _get_compliance_audit_data(request, db, current_user)
            story.extend(_build_compliance_audit_pdf(content, styles))
        
        doc.build(story)
        buffer.seek(0)
        
        filename = f"{request.report_type.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return buffer.getvalue(), filename
        
    except ImportError:
        logger.error("reportlab not installed. Install with: pip install reportlab")
        raise HTTPException(status_code=500, detail="PDF generation not available. Please install reportlab.")


async def generate_excel_report(
    request: ReportRequest,
    db: Session,
    current_user: User
) -> Tuple[bytes, str]:
    """Generate Excel report"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = request.report_type.value.replace('_', ' ').title()
        
        # Header style
        header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title row
        ws['A1'] = request.report_type.value.replace('_', ' ').title() + " Report"
        ws['A1'].font = Font(bold=True, size=16, color="0D9488")
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Metadata rows
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Generated by: {current_user.full_name or current_user.email}"
        if request.start_date or request.end_date:
            date_range = f"{request.start_date or 'Start'} to {request.end_date or 'End'}"
            ws['A4'] = f"Date Range: {date_range}"
        
        # Generate report content based on type
        start_row = 6
        if request.report_type == ReportType.BUILDING_PROGRESS:
            data = await _get_building_progress_data(request, db, current_user)
            start_row = _build_building_progress_excel(ws, data, start_row, header_fill, header_font, border)
        elif request.report_type == ReportType.AGENT_PERFORMANCE:
            data = await _get_agent_performance_data(request, db, current_user)
            start_row = _build_agent_performance_excel(ws, data, start_row, header_fill, header_font, border)
        elif request.report_type == ReportType.INTERACTION_HISTORY:
            data = await _get_interaction_history_data(request, db, current_user)
            start_row = _build_interaction_history_excel(ws, data, start_row, header_fill, header_font, border)
        elif request.report_type == ReportType.COMPLIANCE_AUDIT:
            data = await _get_compliance_audit_data(request, db, current_user)
            start_row = _build_compliance_audit_excel(ws, data, start_row, header_fill, header_font, border)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"{request.report_type.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return buffer.getvalue(), filename
        
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        raise HTTPException(status_code=500, detail="Excel generation not available. Please install openpyxl.")


# Data retrieval functions

async def _get_building_progress_data(request: ReportRequest, db: Session, current_user: User) -> dict:
    """Get building progress data for report"""
    query = db.query(Building).filter(Building.is_deleted == False)
    
    if request.project_id:
        query = query.filter(Building.project_id == request.project_id)
    if request.building_id:
        query = query.filter(Building.building_id == request.building_id)
    
    # Role-based filtering
    if current_user.role == "AGENT":
        query = query.filter(
            or_(
                Building.assigned_agent_id == current_user.user_id,
                Building.assigned_agent_id.is_(None)
            )
        )
    
    buildings = query.order_by(desc(Building.signature_percentage)).all()
    
    # Get project names
    project_ids = {b.project_id for b in buildings}
    projects = {p.project_id: p.project_name for p in db.query(Project).filter(Project.project_id.in_(project_ids)).all()}
    
    return {
        "buildings": [
            {
                "building_name": b.building_name,
                "building_code": b.building_code,
                "project_name": projects.get(b.project_id, "Unknown"),
                "address": b.address,
                "total_units": b.total_units or 0,
                "signature_percentage": float(b.signature_percentage) if b.signature_percentage else 0.0,
                "traffic_light_status": b.traffic_light_status,
                "current_status": b.current_status,
                "created_at": b.created_at.isoformat() if b.created_at else None
            }
            for b in buildings
        ],
        "total_buildings": len(buildings),
        "avg_signature_percentage": sum(float(b.signature_percentage) if b.signature_percentage else 0.0 for b in buildings) / len(buildings) if buildings else 0.0
    }


async def _get_agent_performance_data(request: ReportRequest, db: Session, current_user: User) -> dict:
    """Get agent performance data for report"""
    # Get all agents
    agents_query = db.query(User).filter(User.role == "AGENT", User.is_active == True)
    if request.agent_id:
        agents_query = agents_query.filter(User.user_id == request.agent_id)
    
    agents = agents_query.all()
    
    # Date range filter
    date_filter = None
    if request.start_date or request.end_date:
        conditions = []
        if request.start_date:
            conditions.append(Interaction.interaction_date >= request.start_date)
        if request.end_date:
            conditions.append(Interaction.interaction_date <= request.end_date)
        date_filter = and_(*conditions) if conditions else None
    
    agent_stats = []
    for agent in agents:
        # Count interactions
        interactions_query = db.query(Interaction).filter(Interaction.agent_id == agent.user_id)
        if date_filter:
            interactions_query = interactions_query.filter(date_filter)
        
        total_interactions = interactions_query.count()
        
        # Count signed documents
        signatures_query = db.query(DocumentSignature).join(Owner).filter(
            Owner.assigned_agent_id == agent.user_id,
            DocumentSignature.status == "FINALIZED"
        )
        if date_filter:
            signatures_query = signatures_query.filter(
                DocumentSignature.approved_at >= request.start_date if request.start_date else True,
                DocumentSignature.approved_at <= request.end_date if request.end_date else True
            )
        
        signed_documents = signatures_query.count()
        
        # Count assigned buildings
        assigned_buildings = db.query(Building).filter(
            Building.assigned_agent_id == agent.user_id,
            Building.is_deleted == False
        ).count()
        
        # Count assigned owners
        assigned_owners = db.query(Owner).filter(
            Owner.assigned_agent_id == agent.user_id,
            Owner.is_deleted == False,
            Owner.is_current_owner == True
        ).count()
        
        agent_stats.append({
            "agent_name": agent.full_name or agent.email,
            "email": agent.email,
            "total_interactions": total_interactions,
            "signed_documents": signed_documents,
            "assigned_buildings": assigned_buildings,
            "assigned_owners": assigned_owners,
            "success_rate": (signed_documents / assigned_owners * 100) if assigned_owners > 0 else 0.0
        })
    
    return {
        "agents": agent_stats,
        "total_agents": len(agent_stats),
        "period_start": request.start_date.isoformat() if request.start_date else None,
        "period_end": request.end_date.isoformat() if request.end_date else None
    }


async def _get_interaction_history_data(request: ReportRequest, db: Session, current_user: User) -> dict:
    """Get interaction history data for report"""
    query = db.query(Interaction).join(Owner).filter(Owner.is_deleted == False)
    
    if request.start_date:
        query = query.filter(Interaction.interaction_date >= request.start_date)
    if request.end_date:
        query = query.filter(Interaction.interaction_date <= request.end_date)
    if request.agent_id:
        query = query.filter(Interaction.agent_id == request.agent_id)
    
    # Role-based filtering
    if current_user.role == "AGENT":
        query = query.filter(Interaction.agent_id == current_user.user_id)
    
    interactions = query.order_by(desc(Interaction.interaction_date)).limit(1000).all()
    
    # Get related data
    owner_ids = {i.owner_id for i in interactions}
    agent_ids = {i.agent_id for i in interactions if i.agent_id}
    
    owners = {o.owner_id: o for o in db.query(Owner).filter(Owner.owner_id.in_(owner_ids)).all()}
    agents = {u.user_id: u for u in db.query(User).filter(User.user_id.in_(agent_ids)).all()}
    
    return {
        "interactions": [
            {
                "interaction_date": i.interaction_date.isoformat() if i.interaction_date else None,
                "interaction_type": i.interaction_type,
                "owner_name": owners.get(i.owner_id).full_name if owners.get(i.owner_id) else "Unknown",
                "owner_phone": owners.get(i.owner_id).phone_number if owners.get(i.owner_id) else None,
                "agent_name": agents.get(i.agent_id).full_name if i.agent_id and agents.get(i.agent_id) else "Unknown",
                "summary": i.call_summary,
                "sentiment": i.sentiment,
                "next_action": i.next_action
            }
            for i in interactions
        ],
        "total_interactions": len(interactions)
    }


async def _get_compliance_audit_data(request: ReportRequest, db: Session, current_user: User) -> dict:
    """Get compliance audit data for report"""
    query = db.query(DocumentSignature).join(Owner).filter(Owner.is_deleted == False)
    
    if request.start_date:
        query = query.filter(DocumentSignature.created_at >= datetime.combine(request.start_date, datetime.min.time()))
    if request.end_date:
        query = query.filter(DocumentSignature.created_at <= datetime.combine(request.end_date, datetime.max.time()))
    
    signatures = query.order_by(desc(DocumentSignature.created_at)).limit(1000).all()
    
    # Get related data
    owner_ids = {s.owner_id for s in signatures}
    approver_ids = {s.approved_by_user_id for s in signatures if s.approved_by_user_id}
    
    owners = {o.owner_id: o for o in db.query(Owner).filter(Owner.owner_id.in_(owner_ids)).all()}
    approvers = {u.user_id: u for u in db.query(User).filter(User.user_id.in_(approver_ids)).all()}
    
    return {
        "signatures": [
            {
                "signature_date": s.created_at.isoformat() if s.created_at else None,
                "owner_name": owners.get(s.owner_id).full_name if owners.get(s.owner_id) else "Unknown",
                "status": s.status,
                "approved_at": s.approved_at.isoformat() if s.approved_at else None,
                "approver_name": approvers.get(s.approved_by_user_id).full_name if s.approved_by_user_id and approvers.get(s.approved_by_user_id) else None,
                "approval_reason": s.approval_reason
            }
            for s in signatures
        ],
        "total_signatures": len(signatures)
    }


# PDF building functions

def _build_building_progress_pdf(content: dict, styles) -> List:
    """Build PDF content for building progress report"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    story = []
    
    # Summary
    summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=11, spaceAfter=12)
    story.append(Paragraph(f"Total Buildings: {content['total_buildings']}", summary_style))
    story.append(Paragraph(f"Average Signature Percentage: {content['avg_signature_percentage']:.2f}%", summary_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Table data
    data = [["Building Name", "Project", "Address", "Units", "Signature %", "Status"]]
    
    for building in content['buildings']:
        status_color = {
            "GREEN": "✓",
            "YELLOW": "⚠",
            "RED": "✗"
        }.get(building['traffic_light_status'], "")
        
        data.append([
            building['building_name'],
            building['project_name'],
            building['address'] or "",
            str(building['total_units']),
            f"{building['signature_percentage']:.1f}%",
            f"{status_color} {building['traffic_light_status']}"
        ])
    
    # Create table
    table = Table(data, colWidths=[1.5*inch, 1.2*inch, 1.8*inch, 0.6*inch, 0.8*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D9488')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(table)
    return story


def _build_agent_performance_pdf(content: dict, styles) -> List:
    """Build PDF content for agent performance report"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    
    story = []
    
    # Summary
    summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=11, spaceAfter=12)
    story.append(Paragraph(f"Total Agents: {content['total_agents']}", summary_style))
    if content.get('period_start'):
        story.append(Paragraph(f"Period: {content['period_start']} to {content.get('period_end', 'Present')}", summary_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Table data
    data = [["Agent Name", "Email", "Interactions", "Signed Docs", "Buildings", "Owners", "Success Rate"]]
    
    for agent in content['agents']:
        data.append([
            agent['agent_name'],
            agent['email'],
            str(agent['total_interactions']),
            str(agent['signed_documents']),
            str(agent['assigned_buildings']),
            str(agent['assigned_owners']),
            f"{agent['success_rate']:.1f}%"
        ])
    
    # Create table
    table = Table(data, colWidths=[1.5*inch, 1.5*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.7*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D9488')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(table)
    return story


def _build_interaction_history_pdf(content: dict, styles) -> List:
    """Build PDF content for interaction history report"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    
    story = []
    
    # Summary
    summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=11, spaceAfter=12)
    story.append(Paragraph(f"Total Interactions: {content['total_interactions']}", summary_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Table data
    data = [["Date", "Type", "Owner", "Agent", "Summary", "Sentiment"]]
    
    for interaction in content['interactions'][:100]:  # Limit to 100 for PDF
        data.append([
            interaction['interaction_date'][:10] if interaction['interaction_date'] else "",
            interaction['interaction_type'],
            interaction['owner_name'],
            interaction['agent_name'],
            (interaction['summary'] or "")[:50] + "..." if interaction.get('summary') and len(interaction['summary']) > 50 else (interaction.get('summary') or ""),
            interaction.get('sentiment', '')
        ])
    
    # Create table
    table = Table(data, colWidths=[0.8*inch, 0.8*inch, 1.2*inch, 1*inch, 2*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D9488')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(table)
    return story


def _build_compliance_audit_pdf(content: dict, styles) -> List:
    """Build PDF content for compliance audit report"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    
    story = []
    
    # Summary
    summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=11, spaceAfter=12)
    story.append(Paragraph(f"Total Signatures: {content['total_signatures']}", summary_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Table data
    data = [["Date", "Owner", "Status", "Approved At", "Approver", "Reason"]]
    
    for signature in content['signatures']:
        data.append([
            signature['signature_date'][:10] if signature['signature_date'] else "",
            signature['owner_name'],
            signature['status'],
            signature['approved_at'][:10] if signature['approved_at'] else "",
            signature['approver_name'] or "",
            (signature['approval_reason'] or "")[:40] + "..." if signature.get('approval_reason') and len(signature['approval_reason']) > 40 else (signature.get('approval_reason') or "")
        ])
    
    # Create table
    table = Table(data, colWidths=[0.8*inch, 1.2*inch, 0.8*inch, 0.8*inch, 1*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D9488')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(table)
    return story


# Excel building functions

def _build_building_progress_excel(ws, content: dict, start_row: int, header_fill, header_font, border) -> int:
    """Build Excel content for building progress report"""
    from openpyxl.styles import Alignment
    
    # Headers
    headers = ["Building Name", "Project", "Address", "Units", "Signature %", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for idx, building in enumerate(content['buildings'], 1):
        row = start_row + idx
        ws.cell(row=row, column=1, value=building['building_name']).border = border
        ws.cell(row=row, column=2, value=building['project_name']).border = border
        ws.cell(row=row, column=3, value=building['address'] or "").border = border
        ws.cell(row=row, column=4, value=building['total_units']).border = border
        ws.cell(row=row, column=5, value=f"{building['signature_percentage']:.1f}%").border = border
        ws.cell(row=row, column=6, value=building['traffic_light_status']).border = border
    
    return start_row + len(content['buildings']) + 2


def _build_agent_performance_excel(ws, content: dict, start_row: int, header_fill, header_font, border) -> int:
    """Build Excel content for agent performance report"""
    from openpyxl.styles import Alignment
    
    # Headers
    headers = ["Agent Name", "Email", "Interactions", "Signed Docs", "Buildings", "Owners", "Success Rate"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for idx, agent in enumerate(content['agents'], 1):
        row = start_row + idx
        ws.cell(row=row, column=1, value=agent['agent_name']).border = border
        ws.cell(row=row, column=2, value=agent['email']).border = border
        ws.cell(row=row, column=3, value=agent['total_interactions']).border = border
        ws.cell(row=row, column=4, value=agent['signed_documents']).border = border
        ws.cell(row=row, column=5, value=agent['assigned_buildings']).border = border
        ws.cell(row=row, column=6, value=agent['assigned_owners']).border = border
        ws.cell(row=row, column=7, value=f"{agent['success_rate']:.1f}%").border = border
    
    return start_row + len(content['agents']) + 2


def _build_interaction_history_excel(ws, content: dict, start_row: int, header_fill, header_font, border) -> int:
    """Build Excel content for interaction history report"""
    from openpyxl.styles import Alignment
    
    # Headers
    headers = ["Date", "Type", "Owner", "Agent", "Summary", "Sentiment"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for idx, interaction in enumerate(content['interactions'], 1):
        row = start_row + idx
        ws.cell(row=row, column=1, value=interaction['interaction_date'][:10] if interaction['interaction_date'] else "").border = border
        ws.cell(row=row, column=2, value=interaction['interaction_type']).border = border
        ws.cell(row=row, column=3, value=interaction['owner_name']).border = border
        ws.cell(row=row, column=4, value=interaction['agent_name']).border = border
        ws.cell(row=row, column=5, value=interaction.get('summary', '')).border = border
        ws.cell(row=row, column=6, value=interaction.get('sentiment', '')).border = border
    
    return start_row + len(content['interactions']) + 2


def _build_compliance_audit_excel(ws, content: dict, start_row: int, header_fill, header_font, border) -> int:
    """Build Excel content for compliance audit report"""
    from openpyxl.styles import Alignment
    
    # Headers
    headers = ["Date", "Owner", "Status", "Approved At", "Approver", "Reason"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for idx, signature in enumerate(content['signatures'], 1):
        row = start_row + idx
        ws.cell(row=row, column=1, value=signature['signature_date'][:10] if signature['signature_date'] else "").border = border
        ws.cell(row=row, column=2, value=signature['owner_name']).border = border
        ws.cell(row=row, column=3, value=signature['status']).border = border
        ws.cell(row=row, column=4, value=signature['approved_at'][:10] if signature['approved_at'] else "").border = border
        ws.cell(row=row, column=5, value=signature['approver_name'] or "").border = border
        ws.cell(row=row, column=6, value=signature.get('approval_reason', '')).border = border
    
    return start_row + len(content['signatures']) + 2

